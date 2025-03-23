import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb= xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1,1)

    print(cell.value)
    print(sheet.max_row)

    for row in range(1,sheet.max_row+1):
        print(row)
        
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row, 3)
        coreccted_price = cell.value * 0.9
        coreccted_price_cell = sheet.cell(row,4)
        coreccted_price_cell.value = coreccted_price
        
    values = Reference(sheet,
                min_row=2, 
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

